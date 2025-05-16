import collections
import cvxpy as cp
import numpy as np
import tqdm
import pandas as pd
import re
import shutil
import ast
import nashpy as nash
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from matplotlib.colors import LinearSegmentedColormap
import matplotlib.pyplot as plt

def run_n_rounds_extended_form(D, T, M, A, eps, n_rounds, initial=None, verbose=False):
    # Note: initial should be a tuple of lists specifying indices for initial strategies
    terminals_firm, terminals_work, terminals_dict_f, terminals_dict_w = generate_terminal_seqs(A, n_rounds)
    if n_rounds % 2 == 1:
        terminal_seqs = terminals_firm
    else:
        terminal_seqs = terminals_work

    u_coeffs_worker = generate_u_coefficients(len(terminals_work), terminal_seqs, delta, worker=True)
    u_coeffs_firm = generate_u_coefficients(len(terminals_firm), terminal_seqs, delta, worker=False)

    # generate random pure initial strategies
    if initial is None:
        beta_f, beta_w, n_f, n_w = generate_init_strategies(A, n_rounds, False)
    else:
        beta_f, beta_w, n_f, n_w = generate_init_strategies(A, n_rounds, True, initial[0], initial[1])
    file_path = f'D_{D}_M_{M}_T_{T}_del_{0}_eps_{eps:.0e}_rounds_{n_rounds}_f{tuple(n_f)}_w{tuple(n_w)}.xlsx'

    prev_f = np.array([beta_f])
    prev_w = np.array([beta_w])

    timesteps=0
    for t in tqdm.tqdm(range(T)):
        timesteps+=1
        r_f_t_p_1 = agent_update_memoized(n_rounds, prev_w, A, len(terminals_firm), M, u_coeffs_firm, worker=False)
        r_w_t_p_1 = agent_update_memoized(n_rounds, prev_f, A, len(terminals_work), M, u_coeffs_worker, worker=True)

        prev_f = np.vstack([prev_f, r_f_t_p_1])
        prev_w = np.vstack([prev_w, r_w_t_p_1])

        if check_convergence(prev_f, prev_w, t, convergence_threshold=1e-7):
            print("Converged after", t, "steps")
            is_NE, _, _, _, _ = NE_check(r_f_t_p_1,r_w_t_p_1, A, delta, n_rounds, terminals_firm, terminals_work)
            if is_NE:
                break
    
    if verbose:
        print("----initial strategies----")
        print(f"beta_f: {beta_f}")
        print(f"beta_w: {beta_w}")

    if verbose:
        print("----final convergence----")
        print("--firm--")
        print(r_f_t_p_1)
        print("--worker--")
        print(r_w_t_p_1)
    
    
    # print(get_highest_prob_strat_at_level(1, terminals_firm, r_f_t_p_1))
    r_w_rel = produce_relative_probs(r_w_t_p_1, terminals_work, False, n_rounds)
    r_f_rel = produce_relative_probs(r_f_t_p_1, terminals_firm, True, n_rounds)
    incredible_threats = incredible_threat_check(n_rounds, terminals_firm, r_w_rel, r_f_rel)
    credible_threats = credible_threat_check(delta, r_f_rel ,r_w_rel, A, terminals_dict_w)
    is_NE, strategies, relative_strategies, optimal_strategy_utils, expected_strategy_utils = NE_check(r_f_t_p_1,r_w_t_p_1, A, delta, n_rounds, terminals_firm, terminals_work)
    
    # draw_tree(n_rounds, file_path, A, r_f_rel,r_w_rel, incredible_threats, credible_threats, postfix=postfix)#, show_all_probs=False)
    data_dict = {
        'initial_conditions': (prev_f[0], prev_w[0]),
        'initial_strats': (n_f, n_w),
        'final_strats': strategies, 
        'relative_final_strats': relative_strategies, 
        'optimal_strategy_utilities': optimal_strategy_utils,
        'expected_strategy_utilities_non_negligible': expected_strategy_utils,
        'incredible threats': (incredible_threats, None),
        'credible threats': (credible_threats, None),
        'NE': is_NE,
        'Timesteps':timesteps
    }

    df = pd.DataFrame(data_dict)

    filepath = f'{n_rounds}_round_ne_convergence_files_grid/D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}.xlsx'
    # fig.show()

    try:
        with pd.ExcelWriter(filepath, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
            df.to_excel(writer, sheet_name=f'f {tuple(n_f)}, w {tuple(n_w)}', index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(filepath, mode='w', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'f {tuple(n_f)}, w {tuple(n_w)}', index=False)


    print(f"Coverged to NE: {is_NE}")
    print("--------------------------")
    print()

def run_1_round_normal(A, T=500, M=None, initial=None, reference=None, eps=1e-7):
    if M is None:
        M = np.round(1 / (T**(1/5)),5)
    alpha_f = np.zeros_like(A)
    alpha_w = np.zeros_like(A)
    if reference is not None:
        alpha_f[reference[0]] = 1
        alpha_w[reference[1]] = 1
        ref_str = f'_ref{reference}'
    else:
        ref_str = ''

    beta_f = np.zeros_like(A)
    beta_w = np.zeros_like(A)
    if initial is not None:
        beta_f[initial[0]] = 1
        beta_w[initial[1]] = 1
    filepath = f'{n_rounds}_round_ne_convergence_files_normal_grid/D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}{ref_str}.xlsx'
    prev_f = [beta_f]
    prev_w = [beta_w]

    timesteps = 0
    for t in tqdm.tqdm(range(T)):
        timesteps+=1
        w_f_t_p_1 = agent_update_normal(prev_w, S_i=A, alpha_i=alpha_f, M=M)
        w_w_t_p_1 = agent_update_normal(prev_f, S_i=A, alpha_i=alpha_w, M=M, responder=True)

        prev_f.append(w_f_t_p_1)
        prev_w.append(w_w_t_p_1)

        if check_convergence(prev_f, prev_w, t, window_size=10):
            print(f"Converged after {t} iterations.")
            is_NE, _, _, _ = NE_check_normal(w_f_t_p_1,w_w_t_p_1, A)
            if is_NE:
                break

    is_NE, optimal_strategy_utils, expected_strategy_utils, is_pure = NE_check_normal(w_f_t_p_1,w_w_t_p_1, A)
    print(is_NE, optimal_strategy_utils)
    data_dict = {
        'initial_conditions': (prev_f[0], prev_w[0]),
        'initial_strats': initial,
        'final_strats': (w_f_t_p_1, w_w_t_p_1), 
        'optimal_strategy_utilities': optimal_strategy_utils,
        'expected_strategy_utilities_non_negligible': expected_strategy_utils,
        'NE': is_NE,
        'Pure':is_pure,
        'Timesteps':timesteps
    }

    df = pd.DataFrame(data_dict)
    try:
        with pd.ExcelWriter(filepath, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
            df.to_excel(writer, sheet_name=f'f {tuple([initial[0]])}, w {tuple([initial[1]])}', index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(filepath, mode='w', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'f {tuple([initial[0]])}, w {tuple([initial[1]])}', index=False)

    
def run_1_round_normal_grid(A, D, T=500, M=None, reference=None, eps=1e-7):
    if M is None:
        M = np.round(1 / (T**(1/5)),5)
    alpha_f = np.zeros_like(A)
    alpha_w = np.zeros_like(A)
    if reference is not None:
        alpha_f[reference[0]] = 1
        alpha_w[reference[1]] = 1
        ref_str = f'_ref{reference}'
    else:
        ref_str = ''
    filepath = f'D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}{ref_str}.xlsx'

    for n_f in range(D+1):
        for n_w in range(D+1):
            print(f"candidate strategy {(A[n_w])}, firm strategy {(A[n_f])}")
            try:
                workbook = openpyxl.load_workbook(filepath)
                sheet = workbook[f'f {tuple([n_f])}, w {tuple([n_w])}']
                continue
            except (KeyError,FileNotFoundError):

                beta_f = np.zeros_like(A)
                beta_f[n_f] = 1
                beta_w = np.zeros_like(A)
                beta_w[n_w] = 1

                prev_f = [beta_f]
                prev_w = [beta_w]

                timesteps = 0
                for t in tqdm.tqdm(range(T)):
                    timesteps+=1
                    w_f_t_p_1 = agent_update_normal(prev_w, S_i=A, alpha_i=alpha_f, M=M)
                    w_w_t_p_1 = agent_update_normal(prev_f, S_i=A, alpha_i=alpha_w, M=M, responder=True)

                    prev_f.append(w_f_t_p_1)
                    prev_w.append(w_w_t_p_1)

                    if check_convergence(prev_f, prev_w, t):
                        print(f"Converged after {t} iterations.")
                        is_NE, _, _, _ = NE_check_normal(w_f_t_p_1,w_w_t_p_1, A)
                        if is_NE:
                            break

                is_NE, optimal_strategy_utils, expected_strategy_utils, is_pure = NE_check_normal(w_f_t_p_1,w_w_t_p_1, A)

                data_dict = {
                    'initial_conditions': (prev_f[0], prev_w[0]),
                    'initial_strats': (n_f, n_w),
                    'final_strats': (w_f_t_p_1, w_w_t_p_1), 
                    'optimal_strategy_utilities': optimal_strategy_utils,
                    'expected_strategy_utilities_non_negligible': expected_strategy_utils,
                    'NE': is_NE,
                    'Pure':is_pure,
                    'Timesteps':timesteps
                }

                df = pd.DataFrame(data_dict)
                try:
                    with pd.ExcelWriter(filepath, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
                        df.to_excel(writer, sheet_name=f'f {tuple([n_f])}, w {tuple([n_w])}', index=False)
                except FileNotFoundError:
                    with pd.ExcelWriter(filepath, mode='w', engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=f'f {tuple([n_f])}, w {tuple([n_w])}', index=False)
    generate_imshow_from_sheet(A, pure=True, n_rounds=n_rounds, filepath=filepath, normal=True)

def run_1_round_normal_bar(A, T=500, uniform_firm=True, M=None, reference=None, eps=1e-7):
    if M is None:
        M = np.round(1 / (T**(1/5)),5)

    beta_f = np.zeros(len(A))
    beta_w = np.zeros(len(A))
    if uniform_firm:
        beta_f.fill(1/len(A))
    else:
        beta_w.fill(1/len(A))

    alpha_f = np.zeros_like(A)
    alpha_w = np.zeros_like(A)
    if reference is not None:
        alpha_f[reference[0]] = 1
        alpha_w[reference[1]] = 1
        ref_str = f'_ref{reference}'
    else:
        ref_str = ''
    filepath = f'D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}{ref_str}.xlsx'
    # make_bar_graph(filepath, uniform_firm)
    # exit()
    for n in range(D+1):
        print(f"new strategy {(A[n])}")
        if not uniform_firm:
            beta_f[n] = 1
        else:
            beta_w[n] = 1

        prev_f = [beta_f]
        prev_w = [beta_w]

        timesteps=0
        for t in tqdm.tqdm(range(T)):
            timesteps+=1
            w_f_t_p_1 = agent_update_normal(prev_w, S_i=A, alpha_i=alpha_f, M=M)
            w_w_t_p_1 = agent_update_normal(prev_f, S_i=A, alpha_i=alpha_w, M=M, responder=True)

            prev_f.append(w_f_t_p_1)
            prev_w.append(w_w_t_p_1)

            if check_convergence(prev_f, prev_w, t):
                print(f"Converged after {t} iterations.")
                is_NE, _, _, _ = NE_check_normal(w_f_t_p_1,w_w_t_p_1, A)
                if is_NE:
                    break

        is_NE, optimal_strategy_utils, expected_strategy_utils, is_pure = NE_check_normal(w_f_t_p_1,w_w_t_p_1, A)

        data_dict = {
            'initial_conditions': (prev_f[0], prev_w[0]),
            'initial_strats': ('uniform' if uniform_firm else n, n if uniform_firm else 'uniform'),
            'final_strats': (w_f_t_p_1, w_w_t_p_1), 
            'optimal_strategy_utilities': optimal_strategy_utils,
            'expected_strategy_utilities_non_negligible': expected_strategy_utils,
            'NE': is_NE,
            'Pure':is_pure,
            'Timesteps':timesteps
        }

        df = pd.DataFrame(data_dict)

        
        # fig.show()
        try:
            with pd.ExcelWriter(filepath, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
                df.to_excel(writer, sheet_name=f'idx_{n}', index=False)
        except FileNotFoundError:
            with pd.ExcelWriter(filepath, mode='w', engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=f'idx_{n}', index=False)
        
    make_bar_graph(filepath, uniform_firm)

def make_bar_graph(filepath, uniform_firm=True):
    excel_file = pd.ExcelFile(filepath)
    fpath, fname = filepath.split('/')

    pattern = r"D_(\d+)_M_([0-9.]+)_T_([0-9]+)_del_([0-9.]+)_eps_([0-9eE.-]+)"
    match = re.match(pattern, fname)

    D, M, T, delta, eps = match.groups()
    
    D = int(D)
    M = float(M)
    T = int(T)
    delta = float(delta)
    eps = float(eps)
    # print(D, M, T, delta, eps, n_rounds)

    sheet_names = excel_file.sheet_names
    utils = np.zeros(D+1)
    # print(sheet_names)
    changing = ('Worker','Firm') if uniform_firm else ('Firm','Worker')

    # Loop through each sheet and read the data
    for sheet_name in sheet_names:

        # Pattern to match tuples for f and w
        pattern = r"idx_([0-9.]+)"
        
        match = re.match(pattern, sheet_name)
        n = match.groups()
        n = int(n[0])
        # print(n)
        df = excel_file.parse(sheet_name)
        actual_util = df['expected_strategy_utilities_non_negligible'] 
        is_NE = df['NE'][0]
        # print(actual_util[1])
        if is_NE:
            utils[n] = 1-round(actual_util[0], 4)
        else:
            utils[n] = -1
    print(utils)
    fig, ax = plt.subplots(figsize=(10, 6))
    if n_rounds == 1:
        pattern = r"ref\((.*?)\)"
        # postfix=f'mixed_f({dist_f})_w({dist_w})'
        match = re.search(pattern, fname)
        alpha = None
        if match:
            alpha_str = match.groups()[0]
            # print(alpha_str)
            alpha = tuple(map(int, alpha_str.split(',')))
        title = f"Worker Util Value Given Uniform {changing[1]} Strategy \n(η={M:.2f}, D={D}, T={T}, ɛ={eps:.0e}, reference={alpha}"
    else:
        title = f"Worker Util Value Given Uniform {changing[1]} Strategy \n(η={M:.2f}, D={D}, T={T}, δ={delta:.2f}, ɛ={eps:.0e}"

    fig.suptitle(title)

    # Customize x-axis labels
    x_labels = [f"{n/D:.3f}" for n in range(D+1)]

    plt.bar(x_labels, utils, color='skyblue')
    plt.xticks(rotation=90, ha='right')
    plt.xlabel(f'{changing[0]} Initial Strategy')
    plt.ylabel('Expected Utility for Worker')

    plt.savefig(f'{fpath}/{n_rounds}_round_bar_{fname[0:-5]}.png') 

def agent_update_normal(prev_not_i_strategies, S_i, alpha_i, M, responder= False):

    utility_feedback_vector = [np.float64(0.0) for s in S_i]


    if responder == True: # candidate
        for w_f in prev_not_i_strategies:
            for j, s_c in enumerate(S_i):
                for k, w_f_supp in enumerate(S_i):
                    if w_f_supp >= s_c:
                        utility_feedback_vector[j] += w_f[k] *  w_f_supp
    else: # firm
        for w_c in prev_not_i_strategies:
            for j, s_f in enumerate(S_i):
                for k, w_c_supp in enumerate(S_i):
                    if w_c_supp <= s_f:
                        utility_feedback_vector[j] += w_c[k] *  (1-s_f)

    w_var = cp.Variable(len(S_i))
    
    objective = cp.Maximize(w_var@utility_feedback_vector - cp.norm(w_var-alpha_i, 2)**2/(2*M))

    constraints = [cp.sum(w_var)==1, cp.min(w_var)>=0.0]
    problem = cp.Problem(objective,constraints)

    problem.solve()
    obj_min_strat = [max(np.float64(0.0), w_p.value) for w_p in w_var]

    return obj_min_strat

def NE_check_normal(w_f,w_w, S, eps=1e-7):

    utilities = []
    for i in range(len(S)):
        if i <= np.argmax(w_f):
            utilities.append(sum([w_w[j] for j in range(i+1)])*(1-S[i]))
    
    # print(utilities)
    return (np.argmax(utilities) == np.argmax(w_f), (np.max(utilities), None), (1-A[np.argmax(w_f)], None), 1-np.argmax(w_f) < eps)


def generate_constraints_recursive(n_rounds, r_var, A, sequence="", is_firm=True):
    constraints = []
    n = len(A)

    if n_rounds == 1:
        constraints = [r_var[i]>=0 for i in range(r_var.shape[0])]
        if is_firm:
            constraints.append(cp.sum(r_var[:n]) == 1) # Sum of initial offers = 1
        else:
            for a in A:
                constraints.append(r_var[get_index(f"A{a:.2f}", A)] + r_var[get_index(f"R{a:.2f}", A)] == 1)
        return constraints
    
    if not sequence:  # Base case: root of the game tree
        constraints = [r_var[i]>=0 for i in range(r_var.shape[0])]
        if is_firm:
            constraints.append(cp.sum(r_var[:n]) == 1) # Sum of initial offers = 1
            for a in A:
                constraints.extend(generate_constraints_recursive(n_rounds, r_var, A, f"{a:.2f}", is_firm))
        else:
            for a in A:
                constraints.append(r_var[get_index(f"A{a:.2f}", A)] + 
                                   cp.sum([r_var[get_index(f"R{a:.2f}{b:.2f}", A)] for b in A]) == 1)
                for b in A:
                    constraints.extend(generate_constraints_recursive(n_rounds, r_var, A, f"R{a:.2f}{b:.2f}", is_firm))

    elif len(sequence) == 4 and is_firm:  # Firm's initial offer (a.xx)
        for b in A:
            if n_rounds == 2:
                constraints.append(r_var[get_index(f"{sequence}A{b:.2f}", A)] + r_var[get_index(f"{sequence}R{b:.2f}", A)] == 
                               r_var[get_index(sequence, A)])
            elif n_rounds == 3:
                constraints.append(r_var[get_index(f"{sequence}A{b:.2f}", A)] + 
                                cp.sum([r_var[get_index(f"{sequence}R{b:.2f}{c:.2f}", A)] for c in A]) == 
                                r_var[get_index(sequence, A)])
                constraints.extend(generate_constraints_recursive(n_rounds, r_var, A, f"{sequence}R{b:.2f}", is_firm))

    elif len(sequence) == 9 and not is_firm and n_rounds == 3:  # Worker's counter-offer (Ra.xxb.yy)
        for c in A:
            constraints.append(r_var[get_index(f"{sequence}A{c:.2f}", A)] + 
                               r_var[get_index(f"{sequence}R{c:.2f}", A)] == 
                               r_var[get_index(sequence, A)])

    return constraints


def generate_constraints_iterative(n_rounds, r_var, A, is_firm=True):
    # create summation constraints
    constraints = [r_var[i]>=0 for i in range(r_var.shape[0])]
    n = len(A)

    if is_firm:
        constraints.append(cp.sum(r_var[:n])==1)
        for a in A:
            for b in A:
                if n_rounds == 3:
                    constraints.append(r_var[get_index(f"{a:.2f}A{b:.2f}", A)] + cp.sum([r_var[get_index(f"{a:.2f}R{b:.2f}{c:.2f}", A)] for c in A]) == r_var[get_index(f"{a:.2f}", A)]) # mass on ..Aa_j, Ra_ja_k for each worker offer a_j is an extension of seq (sum to a_i) Note: space is |A|+1 for non terminal
                else:
                    constraints.append(r_var[get_index(f"{a:.2f}A{b:.2f}", A)] + r_var[get_index(f"{a:.2f}R{b:.2f}", A)] == r_var[get_index(f"{a:.2f}", A)]) # mass on ..Aa_j, Ra_ja_k for each worker offer a_j is an extension of seq (sum to a_i) Note: space is |A|+1 for non terminal
        return constraints
    else:
        for a in A:
            constraints.append(r_var[get_index(f"A{a:.2f}", A)] + cp.sum([r_var[get_index(f"R{a:.2f}{b:.2f}", A)] for b in A]) == 1) # A_i, R_ia_j for each firm offer a_i is an extension of the empty sequence for the worker
            if n_rounds == 3:
                for b in A:
                    for c in A:
                        constraints.append(r_var[get_index(f"R{a:.2f}{b:.2f}", A)] == cp.sum(r_var[get_index(f"R{a:.2f}{b:.2f}A{c:.2f}", A)]+r_var[get_index(f"R{a:.2f}{b:.2f}R{c:.2f}", A)])) # mass on ...a_jAa_k and ...a_jRa_k (after firm rejects a_j) sums to mass on Ra_ia_j 
        return constraints
    
def get_payoff(seq, delta):
    # only works from firm perspective
    rnd = seq.count('R')+seq.count('A')
    # print(rnd)
    if len(seq)>=5 and seq[-5:].startswith("A"):
        if len(seq) == 5:
            return (1-float(seq[-4:]))*np.power(delta,rnd-1)
    elif len(seq)>=5 and seq[-5:].startswith("R"):
        return 0
    else:
        # is proposer, find prob other will accept
        rnd = rnd+1
        # print(float(seq[-4:]))
        if len(seq)>0:
            return (float(seq[-4:]))*np.power(delta,rnd-1)
        else:
            return 0

def agent_update_memoized(n_rounds, prev_not_i_strategies, A, tlen, M, u_coefficients, worker=False, ref=None):
    # print(u_coefficients)
    # Vectorized calculation for strategies
    utility_feedback_vector = np.zeros(tlen)
    # print(tlen, len(prev_not_i_strategies[-1]))
    if tlen > len(prev_not_i_strategies[-1]):
        sum_strategies = prev_not_i_strategies.sum(axis=0)
        utility_feedback_vector[:len(sum_strategies)] = u_coefficients[:len(sum_strategies)] * sum_strategies
        # print(worker, utility_feedback_vector)
    else:
        sum_strategies = prev_not_i_strategies.sum(axis=0)
        utility_feedback_vector = u_coefficients * sum_strategies[:len(u_coefficients)]
        # if len(prev_not_i_strategies) >= 1000:
        #     print(sum_strategies[get_index(f'R{0.0:.2f}{0.0:.2f}',A)])
        #     print(worker, f"utility for 0.0 counter offer: {utility_feedback_vector[get_index(f'R{0.0:.2f}{0.0:.2f}',A)]}")
        #     print(sum_strategies[get_index(f'R{0.0:.2f}{0.2:.2f}',A)])
        #     print(worker, f"utility for 0.2 counter offer: {utility_feedback_vector[get_index(f'R{0.0:.2f}{0.2:.2f}',A)]}")
        #     print(sum_strategies[get_index(f'R{0.0:.2f}{0.4:.2f}',A)])
        #     print(worker, f"utility for 0.4 counter offer: {utility_feedback_vector[get_index(f'R{0.0:.2f}{0.4:.2f}',A)]}")

    # utility_feedback_vector = np.zeros(len(terminals))
    # for strategy in prev_not_i_strategies:
    #     utility_feedback_vector += np.array(strategy) * u_coefficients

    r_var = cp.Variable(tlen)
    if n_rounds == 1 and ref is not None:
        objective = cp.Maximize(r_var@utility_feedback_vector - cp.norm(r_var-ref, 2)**2/(2*M))
    else:
        objective = cp.Maximize(r_var@utility_feedback_vector - cp.norm(r_var, 2)**2/(2*M))

    if worker:
        constraints = generate_constraints_recursive(n_rounds, r_var, A, is_firm=False)
    else:
        constraints = generate_constraints_recursive(n_rounds, r_var, A, is_firm=True)

    problem = cp.Problem(objective,constraints)
    problem.solve()

    return [m for m in r_var.value] # accounts for computational errors (cp "correct" up to 1e-8)

def agent_update(n_rounds, prev_not_i_strategies, A, terminals, terminal_seqs, delta, M, worker=False):

    utility_feedback_vector = [0.0 for _ in range(len(terminals))]

    # print(terminal_seqs)
    for r in prev_not_i_strategies: # go through record of other's strategies
        for seq in terminal_seqs:
            rnd = len(seq) // 4
            idx = get_index(seq, A)
            if (worker and rnd % 2 == 1) or (not worker and rnd % 2 == 0):
                utility = float(seq[-4:])*np.power(delta,rnd-1)*r[idx]
            else:
                utility = (1-float(seq[-4:]))*np.power(delta,rnd-1)*r[idx]
            utility_feedback_vector[idx] += utility

    r_var = cp.Variable(len(terminals))
    objective = cp.Maximize(r_var@utility_feedback_vector - cp.norm(r_var, 2)**2/(2*M))

    if worker:
        constraints = generate_constraints_recursive(n_rounds, r_var, A, is_firm=False)
    else:
        constraints = generate_constraints_recursive(n_rounds, r_var, A, is_firm=True)

    problem = cp.Problem(objective,constraints)
    problem.solve()

    return [m for m in r_var.value] # accounts for computational errors (cp "correct" up to 1e-8)

def generate_new_betas(A, n_rounds):
    if n_rounds == 1: 
        return (np.zeros(len(A)), np.zeros(len(A)+len(A)))

    size = len(A)
    for i in range(2,n_rounds+1):
        size+=len(A)**(i)
    
    if n_rounds%2 == 1:
        # firm makes last offer
        return (np.zeros(size), np.zeros(size+len(A)**(n_rounds)))
    else:
        return (np.zeros(size+len(A)**(n_rounds)), np.zeros(size))
    
def generate_terminal_seqs(A, n_rounds):
    # Handles 1, 2 or 3 rounds
    terms_dict_w = collections.defaultdict(list)
    terms_dict_f = collections.defaultdict(list)
    f_terms = []
    w_terms = []
    for a in A:
        f_terms.append(f'{a:.2f}')
        w_terms.append(f'A{a:.2f}')
        terms_dict_w["r1"].append(f'A{a:.2f}')
        terms_dict_f["r1"].append(f'{a:.2f}')
        if n_rounds == 1:
            w_terms.append(f'R{a:.2f}')
    if n_rounds >= 2:
        for a in A:
            for b in A:
                f_terms.append(f'{a:.2f}A{b:.2f}')
                w_terms.append(f'R{a:.2f}{b:.2f}')
                terms_dict_w["r2"].append(f'R{a:.2f}{b:.2f}')
                terms_dict_f["r2"].append(f'R{a:.2f}{b:.2f}')
                if n_rounds == 2:
                    f_terms.append(f'{a:.2f}R{b:.2f}')
                    terms_dict_f["r2"].append(f'{a:.2f}R{b:.2f}')
    if n_rounds == 3:
        for a in A:
            for b in A:
                for c in A:
                    f_terms.append(f'{a:.2f}R{b:.2f}{c:.2f}')
                    w_terms.append(f'R{a:.2f}{b:.2f}A{c:.2f}')
                    w_terms.append(f'R{a:.2f}{b:.2f}R{c:.2f}')
                    terms_dict_f["r3"].append(f'{a:.2f}R{b:.2f}{c:.2f}')
                    terms_dict_w["r3"].append(f'R{a:.2f}{b:.2f}A{c:.2f}')
                    terms_dict_w["r3"].append(f'R{a:.2f}{b:.2f}R{c:.2f}')
    return (f_terms, w_terms, terms_dict_f, terms_dict_w)

def get_index(sequence:str, A):
    n = len(A)
    is_firm = True
    if sequence[0] == 'A' or sequence[0] == 'R':
        is_firm = False
    
    def index_of(value):
        return A.index(float(value))
    
    round = len(sequence)//4
    
    if is_firm:
        if round==1:  # a
            a = float(sequence)
            return index_of(a)
        elif round==2:  # aAb or aRb (2 round)
            if sequence.find('A')>=0:
                a, b = sequence.split('A')
                a = float(a)
                b = float(b)
                return n + index_of(f"{a:.2f}")*n + index_of(f"{b:.2f}")
            else:
                a, b = sequence.split('R')
                a = float(a)
                b = float(b)
                return n + n**2 + index_of(f"{a:.2f}")*n + index_of(f"{b:.2f}")
        elif round==3:  # aRbc
            a, bc = sequence.split('R')
            a = float(a)
            b = float(bc[:4])
            c = float(bc[4:])
            return n + n**2 + index_of(f"{a:.2f}")*(n**2) + index_of(f"{b:.2f}")*n + index_of(f"{c:.2f}")
    else:  # Worker
        if round==1:  # Aa or Ra (1 round)
            a = float(sequence[1:])
            if sequence[0] == 'A':
                # print(sequence)
                return index_of(a)
            return len(A)+index_of(a)
        elif round==2:  # Rab
            ab = sequence[1:]
            a = float(ab[:4])
            b = float(ab[4:])
            return n + index_of(f"{a:.2f}")*n + index_of(f"{b:.2f}")
        elif round==3:  # RabAc RabRc
            abc = sequence[1:]
            a = float(abc[:4])
            b = float(abc[4:8])
            decision = abc[8]
            c = float(abc[9:])
            base_index = n + n**2 + index_of(f"{a:.2f}")*n**2 + index_of(f"{b:.2f}")*n + index_of(f"{c:.2f}")
            return base_index if decision == 'A' else base_index + n**3

def generate_mixed_init_strategies(A, n_rounds, uniform=True, n_f=None, n_w=None, dist_f=None, dist_w=None):
    beta_f, beta_w = generate_new_betas(A, n_rounds)
    # grid over a mix (dist params)
    # vertical axis (firm): the 2 first round offers being mixed over
    # horizontal axis (worker): the counter offer, high/low threshhold

    if uniform:
        beta_f.fill(1/len(A))
        beta_w.fill(1/len(A))
    else:
        # firm
        # mix between 2 first offers
            # cover all offers at least once
            # choose several “mixes” 0.5/0.5, 0.8/0.2, 0.2/0.8
        # mixture over accepting and rejecting
            # choose several “mixes” 0.5/0.5, 0.8/0.2, 0.2/0.8
            # same response strategy for all possible counteroffers

        beta_f[get_index(f"{A[n_f[0]]:.2f}", A)] = dist_f[0]
        beta_f[get_index(f"{A[n_f[1]]:.2f}", A)] = 1-dist_f[0]
        for n in n_f:
            p = beta_f[get_index(f"{A[n]:.2f}", A)]
            for b in A:
                beta_f[get_index(f"{A[n]:.2f}A{b:.2f}", A)] = p*dist_f[1]
                beta_f[get_index(f"{A[n]:.2f}R{b:.2f}", A)] = p-p*dist_f[1]
        
        # worker
        # mix over response strategies
            # 0.5/0.5 accepting + every counter offer
                # same response for all possible offers
            # low offers 0.2 accept 0.8 low counteroffer
                # or low offer 0.2 accept 0.8 high counteroffer
            # high offers 0.8 accept 0.2 high counteroffer
                # or high offers 0.8 accept 0.2 low counter offer
        if dist_w == 'half':
            for a in A:
                beta_w[get_index(f"A{a:.2f}", A)] = 0.5
                beta_w[get_index(f"R{a:.2f}{A[n_w[0]]:.2f}", A)] = 0.5
        else:
            p = 0.2
            # low stored in (n_w[0])
            # high stored in (n_w[1])
            # threshold stored in (n_w[2])
            if dist_w == 'low_high':
                for i,a in enumerate(A):
                    if i <= n_w[2]:
                        beta_w[get_index(f"A{a:.2f}", A)] = p
                        beta_w[get_index(f"R{a:.2f}{A[n_w[0]]:.2f}", A)] = 1-p # counter offer low
                    else:
                        beta_w[get_index(f"A{a:.2f}", A)] = 1-p
                        beta_w[get_index(f"R{a:.2f}{A[n_w[1]]:.2f}", A)] = p # counter offer high
            else:
                for i,a in enumerate(A):
                    if i <= n_w[2]:
                        beta_w[get_index(f"A{a:.2f}", A)] = p
                        beta_w[get_index(f"R{a:.2f}{A[n_w[1]]:.2f}", A)] = 1-p # counter offer high
                    else: 
                        beta_w[get_index(f"A{a:.2f}", A)] = 1-p
                        beta_w[get_index(f"R{a:.2f}{A[n_w[0]]:.2f}", A)] = p # counter offer low

    return (beta_f, beta_w, n_f, n_w)

def generate_init_strategies(A, n_rounds, custom=False, n_f_cust=None, n_w_cust=None):
    # Handles 2,3 rounds
    n_f = []
    n_w = []
    for _ in range(n_rounds):
        n_f.append(np.random.randint(len(A)))
        n_w.append(np.random.randint(len(A)))
    if custom:
        n_f=n_f_cust
        n_w=n_w_cust

    beta_f, beta_w = generate_new_betas(A, n_rounds)

    beta_f[get_index(f"{A[n_f[0]]:.2f}", A)] = 1

    if n_rounds == 1:
        for a in A:
            if a >= A[n_w[0]]:
                beta_w[get_index(f"A{a:.2f}", A)] = 1
            else:
                beta_w[get_index(f"R{a:.2f}", A)] = 1
        return (beta_f, beta_w, n_f, n_w)
    

    for b in A:
        if b <= A[n_f[1]]:
            beta_f[get_index(f"{A[n_f[0]]:.2f}A{b:.2f}", A)] = 1
        else:
            if n_rounds == 2:
                beta_f[get_index(f"{A[n_f[0]]:.2f}R{b:.2f}", A)] = 1
            elif n_rounds == 3:
                beta_f[get_index(f"{A[n_f[0]]:.2f}R{b:.2f}{A[n_f[2]]:.2f}", A)] = 1
    
    # worker
    for a in A:
        if a >= A[n_w[0]]:
            beta_w[get_index(f"A{a:.2f}", A)] = 1
        else:
            beta_w[get_index(f"R{a:.2f}{A[n_w[1]]:.2f}", A)] = 1
        if n_rounds == 3:
            for c in A:
                if c >= A[n_w[2]]:
                    beta_w[get_index(f"R{a:.2f}{A[n_w[1]]:.2f}A{c:.2f}", A)] = 1
                else:
                    beta_w[get_index(f"R{a:.2f}{A[n_w[1]]:.2f}R{c:.2f}", A)] = 1
    return (beta_f, beta_w, n_f, n_w)

def check_convergence(prev_f, prev_w, t, window_size=2, convergence_threshold=1e-7):
    if t >= window_size:
        # f_converged = all(np.linalg.norm(np.array(prev_f[-1]) - np.array(prev_f[-i])) < convergence_threshold for i in range(2, window_size + 2))
        # w_converged = all(np.linalg.norm(np.array(prev_w[-1]) - np.array(prev_w[-i])) < convergence_threshold for i in range(2, window_size + 2))
        f_converged = all(np.linalg.norm(np.array(prev_f[-1]) - np.array(prev_f[-i]), ord=np.inf) < convergence_threshold for i in range(2, window_size + 2))
        w_converged = all(np.linalg.norm(np.array(prev_w[-1]) - np.array(prev_w[-i]), ord=np.inf) < convergence_threshold for i in range(2, window_size + 2))
        # if t>=14000:
        #     print(f"worker strategy change: {np.linalg.norm(np.array(prev_w[-1]) - np.array(prev_w[-2]),ord=np.inf)}")
        #     print(f"firm strategy change: {np.linalg.norm(np.array(prev_f[-1]) - np.array(prev_f[-2]),ord=np.inf)}")
        return f_converged and w_converged
    return False

def get_parent(seq):
    if not seq:
        return ""
    seqr = seq[:-4]
    # print(seq)
    if not seqr[-1].isdigit():
        seqr = seqr[:-1]
    if seqr[0].isdigit() and len(seq) // 4 == 3 and 'R' in seq[-9:]:
        seqr = seq[0:4]

    return seqr

def get_strategies_at_level(round, terminals, is_firm):
    # Firm: a aAb aRbc
    # Worker: Aa Rab RabAc RabRc
    size = 0
    if is_firm:
        if round == 1:
            size = 4
        else: 
            size = 4*round+1
        # print(size)
        return [terminals[i] for i in range(len(terminals)) if (len(terminals[i])==size and terminals[i][0].isdigit())]
    else:
        if round == 1:
            size = 5
        else:
            size = 4*round+(round-1)
        return [terminals[i] for i in range(len(terminals)) if (len(terminals[i])==size and not terminals[i][0].isdigit())]

def convert_terminal(tseq):
    round = len(tseq)//4
    to_firm = False
    if tseq[0] == 'A' or tseq[0] == 'R':
        to_firm = True
    
    if not to_firm:
        if round==1:  # a
            a = float(tseq)
            return [f"A{a:.2f}", f"R{a:.2f}"]
        elif round==2:  # aAb or #aRb
            if tseq.find('A') >= 0:
                a, b = tseq.split('A')
            else:
                a, b = tseq.split('R')
            a = float(a)
            b = float(b)
            return [f"R{a:.2f}{b:.2f}"]
        elif round==3:  # aRbc
            a, bc = tseq.split('R')
            a = float(a)
            b = float(bc[:4])
            c = float(bc[4:])
            return [f"R{a:.2f}{b:.2f}A{c:.2f}", f"R{a:.2f}{b:.2f}R{c:.2f}"]
    else:
        if round==1:  # Aa
            a = float(tseq[1:])
            return [f"{a:.2f}"]
        elif round==2:  # Rab
            ab = tseq[1:]
            a = float(ab[:4])
            b = float(ab[4:])
            return [f"{a:.2f}A{b:.2f}", f"{a:.2f}R{b:.2f}"]
        elif round==3:  # RabAc RabRc
            abc = tseq[1:]
            a = float(abc[:4])
            b = float(abc[4:8])
            c = float(abc[9:])
            return [f"{a:.2f}R{b:.2f}{c:.2f}"]
    print("ERROR")
    return None

def recursive_best_response(cur_round, n_rounds, cur_seq, offer, delta, r_rel, A, is_firm):
    if ((is_firm and cur_round % 2 == 1) or (not is_firm and cur_round % 2 == 0)):
        # proposer perspective
        payoff = delta**(cur_round-1) * (1 - offer)
        proposer = True
        # if is_firm:
        #     print(f"Firm is proposer in {cur_round}, Terminal {cur_seq}")
        # else:
        #     print(f"Worker is proposer in {cur_round}, Terminal {cur_seq}")
    else:
        # responder perspective
        payoff = delta**(cur_round-1) * offer
        proposer = False
        # if is_firm:
        #     print(f"Firm is responder in {cur_round} with potential payoff {payoff}; Terminal {cur_seq}")
        # else:
        #     print(f"Worker is responder in {cur_round} with potential payoff {payoff}; Terminal {cur_seq}")
    
    if cur_round == n_rounds:  # Final round
        if proposer:
            # proposer perspective
            # print(cur_seq)
            converted = convert_terminal(cur_seq)[0]
            # print(converted, r_rel[get_index(converted, A)]*payoff)
            return ([], r_rel[get_index(converted, A)]*payoff)
        else:
            # print(cur_seq, payoff)
            return ([cur_seq], payoff) # should always accept, so utility = payoff
    else:
        if proposer:
            # proposer's perspective
            # parent_seq will include offer (will be a or aRbc for firm, Rab for worker)
            l_seqs = []
            converted = convert_terminal(cur_seq)[0]
            # print(converted)
            
            accept = r_rel[get_index(converted, A)]*payoff
            # print(f'Accept prob of {offer}: ', r_rel[get_index(converted, A)], "utility", r_rel[get_index(converted, A)]*payoff)
            
            reject_u = 0
            
            for o in A:
                # sum over utility for each possible counter offer that can be made
                seq = convert_terminal(cur_seq)[0]
                seq = seq[:-5]+f'R{offer:.2f}{o:.2f}'
                # print(seq)
                prob_offer_o = r_rel[get_index(seq, A)]
                # print("prob offer", prob_offer_o)
                l, utility_o = recursive_best_response(cur_round+1, n_rounds, convert_terminal(seq)[0], o, delta, r_rel, A, is_firm)
                l_seqs+=l
                # print("utility", utility_o)
                reject_u += utility_o*prob_offer_o
            # print('Reject utility: ', reject_u)
            return (l_seqs, accept + reject_u)
        
        else:
            # responder's perspective
            # accept offer --> utility is just payoff for that round

            accept = ([cur_seq], payoff)
            # print('Accept utility: ', accept)
            if not accept[0][-1].find('A')>=0:
                reject_base = 'R'+accept[0][-1]
            else:
                reject_base = accept[0][-1].replace('A','R')
            # print(reject_base)

            # reject offer & counter --> utility depends on optimal counteroffer (recursive computation)
            # get relevant terminals (for (not is_firm)) for child round
            o_star = ([],-1)
            for o in A:
                seq = reject_base + f'{o:.2f}'
                l_seqs, u = recursive_best_response(cur_round+1, n_rounds, seq, o, delta, r_rel, A, is_firm)
                # print(seq, u)
                if u > o_star[1]: # ties: do all possible best responses
                    o_star = (l_seqs+[seq],u)
            best_case_rej_util =  o_star[1]
            # o_star = (o_star[0],best_case_rej_util*prob)
            # print('Best case reject utility: ', best_case_rej_util, "versus payoff", payoff)
            # print("Should reject? ", payoff < best_case_rej_util)
            return accept if payoff > best_case_rej_util else o_star

    
def produce_relative_probs(r, terminals, is_firm, n_rounds):
    rel = np.zeros_like(r)
    if is_firm:
        if n_rounds == 1:
            return r
        for i in range(len(A)):
            rel[i] = r[i]
    else:
        if n_rounds == 3:
            for i in range(len(A)+len(A)**2): 
                rel[i] = r[i]
        else:
            return r
    
    has_parent = n_rounds+1
    if is_firm: 
        has_parent = 2
    if not is_firm and n_rounds == 3:
        has_parent = 3
    
    for rnd in range(n_rounds, has_parent-1, -1):
        strats = get_strategies_at_level(rnd, terminals, is_firm)
        # print(strats)
        for seq in strats:
            parent_seq = get_parent(seq)
            # print(parent_seq, seq)
            parent_prob = r[get_index(parent_seq, A)]
            # print(parent_seq)
            # print(parent_prob)
            seq_idx = get_index(seq,A)
            # if 1-(r[seq_idx]/parent_prob)>eps:
            #     print(seq, r[seq_idx]/parent_prob)
            if parent_prob > 0:
                if parent_prob < 1e-7:
                    rel[seq_idx] = 0
                else:
                    rel[seq_idx] =  min(r[seq_idx]/parent_prob,1)#r[seq_idx]
            else:
                rel[seq_idx] = 0
    # print(rel)
    return rel

def incredible_threat_check(n_rounds, terms, r_rel_worker, r_rel_firm, eps=1e-5):
    # If counter offer not 0, flag firm threat's (reject w some non-negligible probability in last round)

    # Check in first offer branches a<=a* (but counter offer is not 0) (Note: this is only for pure)
        # Make sure agent in last round is approx pure accepting

        # print out strategy of firm where they're making incredible threat (terminal)
            # And the accept/reject probability
        # And print out the actual strategy (a*) and if it's pure

    firm = True
    r = r_rel_firm
    if n_rounds%2 == 1:
        firm = False
        r = r_rel_worker
        # worker accepts/rejects last

    offers = []
    for i,o in enumerate(A):
        if r_rel_firm[i] > eps:
            offers.append((i,o))
    
    a_star = offers[-1][1]
    # print('a_star', a_star)
    incredible = []
    strats = get_strategies_at_level(n_rounds, terms, firm)
    strats = [strat for strat in strats if float(strat[:4])<=a_star and float(strat[-4:])>0]
    # print(strats)
    for seq in strats:
        term = seq
            
        seq_a_idx = get_index(term,A)
        
        # Firm threatens to reject a higher counter offer w some probability
        # given they are accepting a lower counter offer w.p. 1
        if 1-r[seq_a_idx]>eps and (not r[seq_a_idx]<eps):
            # print out strategy of firm where they're making incredible threat (terminal)
                # And the accept/reject probability
            # And print out the actual strategy (a*) and if it's pure
            print(term, "Probability (relative):", f"{r[seq_a_idx]:.4f}") #, "Actual Strategy: " + f"{a_star}", r_rel_firm[get_index(f"{a_star:.2f}", A)])
            seq_to_append = convert_terminal(term)[0]
            seq_to_append = seq_to_append[:-4]+'_'+seq_to_append[-4:]
            incredible.append(seq_to_append)
    return incredible

def credible_threat_check(delta, r_f, r_w, A, terms_dict_w, eps=1e-5):
    # Highlight threats where worker's making lower counter offer (in tree) 
    # Note: threat is off the equilibrium path
    # highlight second level node (first offer)

    # approx pure first offer worker is accepting (approx purely)
        # in all smaller first offer branches, worker should do best response counter offer (approx purely)

    # Credible threat: the worker threatening to reject a lower offer than equilibrium path 
    # and then counter offer something that’s better for them and worse for the firm
    
    eq_path = []
    for i,o in enumerate(A):
        if r_f[i] > eps:
            eq_path.append(f"{o:.2f}")

    credible_threat_seqs = []

    # loop through all offers, and get worker strategies for accepting & rejecting (lower offers than eq path)
    # if worker threatens to reject an offer lower than eq path with some probability, and then counter
    # something better for them and worse for the firm (check utility > or < that in eq path), it's a credible threat.

    if float(eq_path[0]) > 0:
        # print('eq path round 1', eq_path[0])
        for strat in terms_dict_w["r1"]:
            if float(strat[1:5]) < float(eq_path[0]):
                # check if worker threatening to reject
                prob_accept = r_w[get_index(strat, A)]
                if prob_accept < eps:
                    # check if counter offer yields utility > or < that in eq path
                    eq_path_util = (1-float(eq_path[0]), float(eq_path[0])) # (firm_util, work_util)
                    for i,o in enumerate(A):
                        counter_seq = 'R'+strat[1:5]+f"{o:.2f}"
                        if r_w[get_index(counter_seq,A)] > eps:
                            # print('worker strat non-eq path round 2', counter_seq)
                            counter_util = (delta*o, delta*(1-o))
                            # print(eq_path_util, counter_util)
                            # credible if the worker is getting the highest expected payoff they can
                            # threat if the firm gets less than other branch (they get more)
                            if i <= 1 and (counter_util[0] < eq_path_util[0]):
                                credible_threat_seqs.append(strat[1:5])

    return credible_threat_seqs

def get_highest_prob_strat_at_level(level, terms, r_f):
    # note: only works for level 1 right now
    strats_at_level = get_strategies_at_level(level, terminals=terms, is_firm=True)
    # print([r_f[get_index(seq, A, is_firm=True)] for seq in strats_at_level])
    return strats_at_level[np.argmax([r_f[get_index(seq, A)] for seq in strats_at_level])]

import igraph as ig
from igraph import Graph, EdgeSeq
import plotly.graph_objects as go
import os
import posixpath as path
# import gravis as gv

def generate_1_round_graph(r_f_rel, r_w_rel, A):
    graph = Graph(directed=True)
    vertices = ["ROOT"]
    for a in A:
        vertices.append(f"{a:.2f}")
        vertices.append(f"A{a:.2f}")
        vertices.append(f"R{a:.2f}")

    graph.add_vertices(vertices)
    for a in A:
        graph.add_edge(f"ROOT", f"{a:.2f}", **{'probability': f"{r_f_rel[get_index(f'{a:.2f}', A)]:.5f}"})
        graph.add_edge(f"{a:.2f}", f"A{a:.2f}", **{'probability': f"{r_w_rel[get_index(f'A{a:.2f}', A)]:.5f}", 'payoff (firm, worker)':(1-a, a)})
        graph.add_edge(f"{a:.2f}", f"R{a:.2f}", **{'probability': f"{r_w_rel[get_index(f'R{a:.2f}', A)]:.5f}", 'payoff (firm, worker)':(0, 0)})
    return graph, vertices

def generate_2_round_graph(r_f_rel, r_w_rel, A):
    graph = Graph(directed=True)
    vertices = ["ROOT"]
    for a in A:
        vertices.append(f"{a:.2f}")
        vertices.append(f"A{a:.2f}")
        for b in A:
            vertices.append(f"R{a:.2f}_{b:.2f}")
            vertices.append(f"{a:.2f}A{b:.2f}")
            vertices.append(f"{a:.2f}R{b:.2f}")

    graph.add_vertices(vertices)
    for a in A:
        graph.add_edge(f"ROOT", f"{a:.2f}", **{'probability': f"{r_f_rel[get_index(f'{a:.2f}', A)]:.4e}"})
        graph.add_edge(f"{a:.2f}", f"A{a:.2f}", **{'probability': f"{r_w_rel[get_index(f'A{a:.2f}', A)]:.4e}", 'payoff (firm, worker)':(1-a, a)})
        for b in A:
            graph.add_edge(f"{a:.2f}", f"R{a:.2f}_{b:.2f}", **{'probability': f"{r_w_rel[get_index(f'R{a:.2f}{b:.2f}', A)]:.4e}"})
            graph.add_edge(f"R{a:.2f}_{b:.2f}", f"{a:.2f}A{b:.2f}", **{'probability': f"{r_f_rel[get_index(f'{a:.2f}A{b:.2f}', A)]:.4e}", 'payoff (firm, worker)':(b*delta, (1-b)*delta)})
            graph.add_edge(f"R{a:.2f}_{b:.2f}", f"{a:.2f}R{b:.2f}", **{'probability': f"{r_f_rel[get_index(f'{a:.2f}R{b:.2f}', A)]:.4e}", 'payoff (firm, worker)':(0, 0)})
    return graph, vertices

def generate_3_round_graph(r_f_rel, r_w_rel, A):
    graph = Graph(directed=True)
    vertices = ["ROOT"]
    for a in A:
        vertices.append(f"{a:.2f}")
        vertices.append(f"A{a:.2f}")
        for b in A:
            vertices.append(f"R{a:.2f}_{b:.2f}")
            vertices.append(f"{a:.2f}A{b:.2f}")
            for c in A:
                vertices.append(f"{a:.2f}R{b:.2f}_{c:.2f}")
                vertices.append(f"R{a:.2f}_{b:.2f}A{c:.2f}")
                vertices.append(f"R{a:.2f}_{b:.2f}R{c:.2f}")

    graph.add_vertices(vertices)
    for a in A:
        graph.add_edge(f"ROOT", f"{a:.2f}", **{'probability': f"{r_f_rel[get_index(f'{a:.2f}', A)]:.4e}"})
        graph.add_edge(f"{a:.2f}", f"A{a:.2f}", **{'probability': f"{r_w_rel[get_index(f'A{a:.2f}', A)]:.4e}", 'payoff (firm, worker)':(1-a, a)})
        for b in A:
            graph.add_edge(f"{a:.2f}", f"R{a:.2f}_{b:.2f}", **{'probability': f"{r_w_rel[get_index(f'R{a:.2f}{b:.2f}', A)]:.4e}"})
            graph.add_edge(f"R{a:.2f}_{b:.2f}", f"{a:.2f}A{b:.2f}", **{'probability': f"{r_f_rel[get_index(f'{a:.2f}A{b:.2f}', A)]:.4e}", 'payoff (firm, worker)':(b*delta, (1-b)*delta)})
            for c in A:
                graph.add_edge(f"R{a:.2f}_{b:.2f}", f"{a:.2f}R{b:.2f}_{c:.2f}", **{'probability': f"{r_f_rel[get_index(f'{a:.2f}R{b:.2f}{c:.2f}', A)]:.4e}"})
                graph.add_edge(f"{a:.2f}R{b:.2f}_{c:.2f}", f"R{a:.2f}_{b:.2f}A{c:.2f}", **{'probability': f"{r_w_rel[get_index(f'R{a:.2f}{b:.2f}A{c:.2f}', A)]:.4e}", 'payoff (firm, worker)':((1-c)*delta**2, c*delta**2)})
                graph.add_edge(f"{a:.2f}R{b:.2f}_{c:.2f}", f"R{a:.2f}_{b:.2f}R{c:.2f}", **{'probability': f"{r_w_rel[get_index(f'R{a:.2f}{b:.2f}R{c:.2f}', A)]:.4e}", 'payoff (firm, worker)':(0,0)})
    return graph, vertices

def draw_tree(n_rounds, index, A, r_f_rel, r_w_rel, incredible_threats, credible_threats, show_all_probs=True, postfix='show'):
    if n_rounds == 2:
        graph, vertices = generate_2_round_graph(r_f_rel, r_w_rel, A)
    elif n_rounds == 1:
        graph, vertices = generate_1_round_graph(r_f_rel, r_w_rel, A)
    else:
        graph, vertices = generate_3_round_graph(r_f_rel, r_w_rel, A)
    num_v = len(vertices)
    
    lay = graph.layout('rt')
    position = {k: lay[k] for k in range(num_v)}
    Y = [lay[k][1] for k in range(num_v)]
    M = max(Y)

    # Extract edges
    es = EdgeSeq(graph)
    E = [e.tuple for e in graph.es]

    # Prepare coordinates for edges and nodes
    Xn = [position[k][0] for k in range(num_v)]
    Yn = [2*M-position[k][1] for k in range(num_v)]
    Xe = []
    Ye = []
    for edge in E:
        Xe+=[position[edge[0]][0],position[edge[1]][0], None]
        Ye+=[2*M-position[edge[0]][1],2*M-position[edge[1]][1], None]

    # Create Plotly traces
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=Xe,
                    y=Ye,
                    mode='lines',
                    line=dict(color='rgb(210,210,210)', width=1),
                    hoverinfo='none'
                    ))
    
    # Add node labels
    node_labels = vertices

    # Add this before creating the Scatter plot for nodes
    node_colors = []
    for node in vertices:
        if node in incredible_threats:
            node_colors.append("#FF0000")  # Red
        elif node in credible_threats:  
            node_colors.append("#FFA500")  # Orange
        else:
            node_colors.append("#6175c1")  # Default blue  # Green 00FF00 
    
    fig.add_trace(go.Scatter(x=Xn,
                    y=Yn,
                    mode='markers',
                    name='bla',
                    marker=dict(symbol='circle-dot',
                                    size=18,
                                    color= node_colors,#'#6175c1',    #'#DB4551',
                                    line=dict(color='rgb(50,50,50)', width=1)
                                    ),
                    text=node_labels,
                    hoverinfo='text',
                    opacity=0.8
                    ))

    # Create text inside the circle via annotations
    def make_annotations(pos, text, font_size=10, font_color='rgb(250,250,250)'):
        L=len(pos)
        if len(text)!=L:
            raise ValueError('The lists pos and text must have the same len')
        annotations = []
        for k in range(L):
            annotations.append(
                dict(
                    text=text[k], # or replace labels with a different list for the text within the circle
                    x=pos[k][0], y=2*M-position[k][1],
                    xref='x1', yref='y1',
                    font=dict(color=font_color, size=font_size),
                    showarrow=False)
            )
        return annotations
    
    # Add edge labels (probabilities and payoffs)
    edge_labels = []
    for e in graph.es:
        label = f"Pr: {e['probability'] if (show_all_probs or (not show_all_probs and float(e['probability'])>1e-5)) else '≈0'}"
        if 'payoff (firm, worker)' in e.attributes() and e['payoff (firm, worker)']:
            label += f"\nPayoff: {e['payoff (firm, worker)']}"
        edge_labels.append(label)

    # Add edge labels to the plot
    for i, edge in enumerate(E):
        x = (position[edge[0]][0] + position[edge[1]][0]) / 2
        y = (2*M-position[edge[0]][1] + 2*M-position[edge[1]][1]) / 2
        fig.add_annotation(x=x, y=y, text=edge_labels[i], textangle=-45,showarrow=False, font_size=8)

    # Add axis specifications and create the layout
    axis = dict(showline=False, # hide axis line, grid, ticklabels and  title
                zeroline=False,
                showgrid=False,
                showticklabels=False,
                )

    fig.update_layout(title= f'Tree Layout of {n_rounds} Round Game',
                annotations=make_annotations(position, node_labels),
                font_size=12,
                showlegend=False,
                xaxis=axis,
                yaxis=axis,
                margin=dict(l=40, r=40, b=85, t=100),
                hovermode='closest',
                plot_bgcolor='rgb(248,248,248)'
                )
    # gv.save(graph, f"2_round_tree_{i}.html")
    file_path = f"{n_rounds}_round_plots_{postfix}/{index}.html"
    # fig.show()
    absolute_path = os.path.abspath(file_path)
    directory = os.path.dirname(absolute_path)
    os.makedirs(directory, exist_ok=True) 
    fig.write_html(file=absolute_path)

def NE_check(r_f, r_w, A, delta, n_rounds, terminals_firm, terminals_work, eps=1e-7): # change default eps
    # Best utility will be within epsilon of the current strategy if in NE
        # Assuming other plays with probability 1
    # At last time step, calculate U_F^T times r_f >= Fopt - epsilon
        # Fopt is utility calculated from best response for firm
    # At last time step, calculate U_W^T times r_w >= Wopt - epsilon
        # Wopt is utility calculated from best response for worker
    # IF BOTH TRUE, IN NE!

    # Created a function to yield total utility given both players (pure-ish) strategies
    # Set up best response to output pure strategy on best response indices to get comparable "best" strategy
    # Firm utility: f^T dot Uf  Worker utility: w^T dot Uw
    # print(r_f, r_w)
    r_f_rel = produce_relative_probs(r_f, terminals_firm, is_firm=True, n_rounds=n_rounds)
    r_w_rel = produce_relative_probs(r_w, terminals_work, is_firm=False, n_rounds=n_rounds)

    terminals = terminals_firm
    convert_firm = False
    if n_rounds % 2 == 0: # use worker's
        convert_firm = True
        terminals = terminals_work

    opt_f = []
    opt_w = []
    first_round_utils = []
    for seq in terminals_firm[:len(A)]:
        round = len(seq)//4

        l_seqs, _ = recursive_best_response(round, n_rounds, convert_terminal(seq)[0], float(seq[-4:]), delta, r_f_rel, A, is_firm=False)
        opt_w+=l_seqs

        l_seqs, firm_best_response_util = recursive_best_response(round, n_rounds, seq, float(seq[-4:]), delta, r_w_rel, A, is_firm=True)
        opt_f+=l_seqs
        first_round_utils.append(firm_best_response_util)
    opt_w.sort(key=lambda s: len(s))
    opt_f.sort(key=lambda s: len(s))

    print(opt_w)
    print(opt_f)
    
    first_round_best_idx = np.argmax(first_round_utils)
    first_round_offer_seq = f"{A[first_round_best_idx]:.2f}"

    # print(first_round_offer_seq)
    # Go top-down to collect best response down NE path (only put 1's in branches where you have 1 at parent)
    # Track down to terminal sequence
    # Utility feedback vector generated by the worker at last time step should be dot producted for ideal firm's (pure opt strat)
    # print(opt_f)
    # print(opt_w)
    work_pure_opt = np.zeros_like(r_w)
    firm_pure_opt = np.zeros_like(r_f)

    firm_pure_opt[get_index(first_round_offer_seq,A)] = 1
    if n_rounds == 1:
        for s in opt_w:
            work_pure_opt[get_index(s,A)] = 1
    # print(firm_pure_opt, work_pure_opt)
    # for i,_ in enumerate(A):
    #     firm_pure_opt[i] = r_f[i]

    # Go top-down, only put 1s on parents w ones
    # TODO: fix for third case
    if n_rounds == 2:

        second_round_offer_seq_w = {}
        for s in opt_w:
            if len(s) // 4 <= 2:
                work_pure_opt[get_index(s,A)] = 1
                second_round_offer_seq_w[s[1:5]] = s[5:9] if len(s)>5 else s
            elif s[5:9] == second_round_offer_seq_w[s[1:5]]:
                work_pure_opt[get_index(s,A)] = 1
        for s in opt_f:
            if s[0:4] == first_round_offer_seq:
                firm_pure_opt[get_index(s,A)] = 1

    # print(second_round_offer_seq_w)

    utility_feedback_vector_work = [0.0 for _ in range(len(terminals_work))]
    utility_feedback_vector_firm = [0.0 for _ in range(len(terminals_firm))]

    for seq in terminals:
        if not convert_firm: # convert worker
            seq = convert_terminal(seq)[0]
        rnd = len(seq) // 4
        idx = get_index(seq, A)
        if rnd % 2 == 1:
            utility = float(seq[-4:])*np.power(delta,rnd-1)*r_f[idx]
        else:
            utility = (1-float(seq[-4:]))*np.power(delta,rnd-1)*r_f[idx]
        utility_feedback_vector_work[idx] += utility

    for seq in terminals:
        if convert_firm:
            seq = convert_terminal(seq)[0]
        rnd = len(seq) // 4
        idx = get_index(seq, A)
        if rnd % 2 == 0:
            utility = float(seq[-4:])*np.power(delta,rnd-1)*r_w[idx]
        else:
            utility = (1-float(seq[-4:]))*np.power(delta,rnd-1)*r_w[idx]
        utility_feedback_vector_firm[idx] += utility

    # Verify worker's/firm's strategy r_w given the firm's/worker's fixed (pure) choices
        # Expected Utility for Worker = r_w dot utility_feedback_vector_work
        # Expected Utility for Firm = r_f dot utility_feedback_vector_firm
    act_util_work = np.dot(r_w, utility_feedback_vector_work)
    act_util_firm = np.dot(r_f, utility_feedback_vector_firm)

    opt_util_work = np.dot(work_pure_opt, utility_feedback_vector_work)
    opt_util_firm = np.dot(firm_pure_opt, utility_feedback_vector_firm)

    print("Actual Util: ", (act_util_firm, act_util_work))
    print("Nash Util: ", (opt_util_firm, opt_util_work))

    NE = False
    result_tuple = tuple(abs(x - y) for x, y in zip((act_util_firm, act_util_work), (opt_util_firm, opt_util_work)))
    if result_tuple < (eps,eps):
        NE = True

    return NE, (r_f, r_w), (r_f_rel, r_w_rel), (opt_util_firm, opt_util_work), (act_util_firm, act_util_work)

def generate_u_coefficients(tlen, terminal_seqs, delta, worker):
    # Precompute coefficients (unchanging part)
    u_coeff_vector = np.zeros(tlen)
    for seq in terminal_seqs:
        rnd = len(seq) // 4
        idx = get_index(seq, A)
        
        # Compute coefficient once per sequence
        if (worker and rnd % 2 == 1) or (not worker and rnd % 2 == 0):
            coeff = float(seq[-4:]) * (delta ** (rnd-1))
        else:
            coeff = (1 - float(seq[-4:])) * (delta ** (rnd-1))
        
        u_coeff_vector[idx] += coeff  # Accumulate in case of duplicate indices
    return u_coeff_vector

def run_1_round_EF_grid(D, T, M, A, eps, reference=(None,None), verbose=False):
    n_rounds = 1
    terminals_firm, terminals_work, _, _ = generate_terminal_seqs(A, n_rounds)
    if n_rounds % 2 == 1:
        terminal_seqs = terminals_firm
    else:
        terminal_seqs = terminals_work

    u_coeffs_worker = generate_u_coefficients(len(terminals_work), terminal_seqs, delta, worker=True)
    u_coeffs_firm = generate_u_coefficients(len(terminals_firm), terminal_seqs, delta, worker=False)
    ref_str = ''
    if reference != (None, None):
        ref_str = f'_ref{reference}'
    alpha = (None, None)
    if reference != (None, None):
        alpha = ([1 if i == reference[0] else 0 for i in range(len(terminals_firm))], [1 if i == reference[1] else 0 for i in range(len(terminals_work))])

    for n_f1 in range(0,D+1):
        for n_w1 in range(0,D+1):
            print(f"Firm strategy indices {(n_f1)}, Worker strategy indices {(n_w1)}")
            # generate random pure initial strategies
            beta_f, beta_w, n_f, n_w = generate_init_strategies(A, n_rounds, True, [n_f1], [n_w1])
            file_path = f'D_{D}_M_{M}_T_{T}_del_{0}_eps_{eps:.0e}_rounds_{n_rounds}_f{tuple(n_f)}_w{tuple(n_w)}'+ref_str
            postfix='grid'
            # print(beta_f, beta_w)

            # if os.path.exists(f'{n_rounds}_round_plots_{postfix}/{file_path}.html'):
            #     continue
            prev_f = np.array([beta_f])
            prev_w = np.array([beta_w])

            timesteps=0
            for t in tqdm.tqdm(range(T)):
                timesteps+=1
                r_f_t_p_1 = agent_update_memoized(n_rounds, prev_w, A, len(terminals_firm), M, u_coeffs_firm, worker=False, ref=alpha[0])
                r_w_t_p_1 = agent_update_memoized(n_rounds, prev_f, A, len(terminals_work), M, u_coeffs_worker, worker=True, ref=alpha[1])

                prev_f = np.vstack([prev_f, r_f_t_p_1])
                prev_w = np.vstack([prev_w, r_w_t_p_1])

                if check_convergence(prev_f, prev_w, t, convergence_threshold=1e-7):
                    print("Converged after", t, "steps")
                    is_NE, _, _, _, _ = NE_check(r_f_t_p_1,r_w_t_p_1, A, delta, n_rounds, terminals_firm, terminals_work)
                    if is_NE:
                        break
            
            if verbose:
                print("----initial strategies----")
                print(f"beta_f: {beta_f}")
                print(f"beta_w: {beta_w}")

            if verbose:
                print("----final convergence----")
                print("--firm--")
                print(r_f_t_p_1)
                print("--worker--")
                print(r_w_t_p_1)
            
            
            is_NE, strategies, relative_strategies, optimal_strategy_utils, expected_strategy_utils = NE_check(r_f_t_p_1,r_w_t_p_1, A, delta, n_rounds, terminals_firm, terminals_work)
            
            # draw_tree(n_rounds, file_path, A, r_f_rel,r_w_rel, incredible_threats, credible_threats, postfix=postfix)#, show_all_probs=False)
            data_dict = {
                'initial_conditions': (prev_f[0], prev_w[0]),
                'initial_strats': (n_f, n_w),
                'final_strats': strategies, 
                'relative_final_strats': relative_strategies, 
                'optimal_strategy_utilities': optimal_strategy_utils,
                'expected_strategy_utilities_non_negligible': expected_strategy_utils,
                # 'incredible threats': (incredible_threats, None),
                # 'credible threats': (credible_threats, None),
                'NE': is_NE,
                'Timesteps': timesteps
            }

            df = pd.DataFrame(data_dict)

            filepath = f'{n_rounds}_round_ne_convergence_files_grid/D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}{ref_str}.xlsx'
            # fig.show()

            try:
                with pd.ExcelWriter(filepath, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
                    df.to_excel(writer, sheet_name=f'f {tuple(n_f)}, w {tuple(n_w)}', index=False)
            except FileNotFoundError:
                with pd.ExcelWriter(filepath, mode='w', engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=f'f {tuple(n_f)}, w {tuple(n_w)}', index=False)

            if verbose:
                print(f"Coverged to NE: {is_NE}")
                print("--------------------------")
                print()

def run_2_round_grid_search(D, T, M, A, delta, eps, verbose=False):
    n_rounds = 2
    terminals_firm, terminals_work, _, terminals_dict_w = generate_terminal_seqs(A, n_rounds)
    if n_rounds % 2 == 1:
        terminal_seqs = terminals_firm
    else:
        terminal_seqs = terminals_work

    u_coeffs_worker = generate_u_coefficients(len(terminals_work), terminal_seqs, delta, worker=True)
    u_coeffs_firm = generate_u_coefficients(len(terminals_firm), terminal_seqs, delta, worker=False)

    for n_f1 in range(0,D+1):
        for n_f2 in range(0,D+1):
            # payoffs = []
            for n_w1 in range(0,D+1):
                for n_w2 in range(0,D+1):
                    print(f"Firm strategy indices {(n_f1, n_f2)}, Worker strategy indices {(n_w1, n_w2)}")
                    # generate random pure initial strategies
                    beta_f, beta_w, n_f, n_w = generate_init_strategies(A, n_rounds, True, [n_f1, n_f2], [n_w1, n_w2])
                    index = f'D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}'
                    inits=f'f{tuple(n_f)}_w{tuple(n_w)}'
                    file_path = index+'/'+inits
                    postfix='grid'
                    
                    
                    try:
                        workbook = openpyxl.load_workbook(f'D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}.xlsx')
                        sheet = workbook[f'f {tuple(n_f)}, w {tuple(n_w)}']
                        continue
                    except (KeyError,FileNotFoundError):
                        # if os.path.exists(f'{n_rounds}_round_plots_{postfix}/{file_path}.html'):
                        #     continue
                        prev_f = np.array([beta_f])
                        prev_w = np.array([beta_w])

                        timesteps=0
                        for t in tqdm.tqdm(range(T)):
                            timesteps+=1
                            r_f_t_p_1 = agent_update_memoized(n_rounds, prev_w, A, len(terminals_firm), M, u_coeffs_firm, worker=False)
                            r_w_t_p_1 = agent_update_memoized(n_rounds, prev_f, A, len(terminals_work), M, u_coeffs_worker, worker=True)

                            prev_f = np.vstack([prev_f, r_f_t_p_1])
                            prev_w = np.vstack([prev_w, r_w_t_p_1])

                            if check_convergence(prev_f, prev_w, t, convergence_threshold=eps):
                                print("Converged after", t, "steps")
                                is_NE, _, _, _, _ = NE_check(r_f_t_p_1,r_w_t_p_1, A, delta, n_rounds, terminals_firm, terminals_work)
                                if is_NE:
                                    break
                        
                        if verbose:
                            print("----initial strategies----")
                            print(f"beta_f: {beta_f}")
                            print(f"beta_w: {beta_w}")

                        if verbose:
                            print("----final convergence----")
                            print("--firm--")
                            print(r_f_t_p_1)
                            print("--worker--")
                            print(r_w_t_p_1)
                        
                        # print(get_highest_prob_strat_at_level(1, terminals_firm, r_f_t_p_1))
                        r_w_rel = produce_relative_probs(r_w_t_p_1, terminals_work, False, n_rounds)
                        r_f_rel = produce_relative_probs(r_f_t_p_1, terminals_firm, True, n_rounds)
                        incredible_threats = incredible_threat_check(n_rounds, terminals_firm, r_w_rel, r_f_rel)
                        credible_threats = credible_threat_check(delta, r_f_rel ,r_w_rel, A, terminals_dict_w)
                        is_NE, strategies, relative_strategies, optimal_strategy_utils, expected_strategy_utils = NE_check(r_f_t_p_1,r_w_t_p_1, A, delta, n_rounds, terminals_firm, terminals_work)
                        
                        # draw_tree(n_rounds, file_path, A, r_f_rel,r_w_rel, incredible_threats, credible_threats, postfix=postfix)#, show_all_probs=False)
                        data_dict = {
                            'initial_conditions': (prev_f[0], prev_w[0]),
                            'initial_strats': (n_f, n_w),
                            'final_strats': strategies, 
                            'relative_final_strats': relative_strategies, 
                            'optimal_strategy_utilities': optimal_strategy_utils,
                            'expected_strategy_utilities_non_negligible': expected_strategy_utils,
                            'incredible threats': (incredible_threats, None),
                            'credible threats': (credible_threats, None),
                            'NE': is_NE,
                            'Timesteps':timesteps
                        }

                        df = pd.DataFrame(data_dict)

                        filepath = f'D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}.xlsx'
                        # fig.show()

                        try:
                            with pd.ExcelWriter(filepath, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
                                df.to_excel(writer, sheet_name=f'f {tuple(n_f)}, w {tuple(n_w)}', index=False)
                        except FileNotFoundError:
                            with pd.ExcelWriter(filepath, mode='w', engine='openpyxl') as writer:
                                df.to_excel(writer, sheet_name=f'f {tuple(n_f)}, w {tuple(n_w)}', index=False)

                        if verbose:
                            print(f"Coverged to NE: {is_NE}")
                            print("--------------------------")
                            print()
    generate_imshow_from_sheet(A, pure=True, n_rounds=n_rounds, filepath=filepath, normal=False)

def run_2_round_mixed_init_grid_search(D, T, M, A, delta, eps, verbose=False):
    n_rounds = 2
    terminals_firm, terminals_work, _, terminals_dict_w = generate_terminal_seqs(A, n_rounds)
    if n_rounds % 2 == 1:
        terminal_seqs = terminals_firm
    else:
        terminal_seqs = terminals_work

    u_coeffs_worker = generate_u_coefficients(len(terminals_work), terminal_seqs, delta, worker=True)
    u_coeffs_firm = generate_u_coefficients(len(terminals_firm), terminal_seqs, delta, worker=False)

    firm_mix_options = [0.5, 0.8, 0.2]
    worker_mix_options = ['high_low'] # 'half', 'low_high', 
    # loop through all possible mixes
    for dist_w in worker_mix_options:
        for first in firm_mix_options:
            for second in firm_mix_options:
                dist_f = [first, second]
                # loop through all possible n_f and n_w
                for i in range(len(A)):
                    for j in range(i+1,len(A)):
                        # permutation of first two offers to mix over (firm)
                        n_f = [i,j]
                        for k in range(len(A)//2):
                            for l in range(len(A)//2, len(A)):
                                # permutation of low and high offers to mix over (work)
                                n_w = [k, l, len(A)//2-1] # threshold is just half (corrected for 0 indexing)
                                beta_f, beta_w, n_f, n_w = generate_mixed_init_strategies(A, n_rounds, False, n_f, n_w, dist_f, dist_w)

                                file_path = f'D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}_f{tuple(n_f)}_w{tuple(n_w)}'
                                
                                prev_f = np.array([beta_f])
                                prev_w = np.array([beta_w])

                                timesteps=0
                                for t in tqdm.tqdm(range(T)):
                                    timesteps+=1
                                    r_f_t_p_1 = agent_update_memoized(n_rounds, prev_w, A, len(terminals_firm), M, u_coeffs_firm, worker=False)
                                    r_w_t_p_1 = agent_update_memoized(n_rounds, prev_f, A, len(terminals_work), M, u_coeffs_worker, worker=True)

                                    prev_f = np.vstack([prev_f, r_f_t_p_1])
                                    prev_w = np.vstack([prev_w, r_w_t_p_1])

                                    if check_convergence(prev_f, prev_w, t,convergence_threshold=1e-7):
                                        print("Converged after", t, "steps")
                                        is_NE, _, _, _, _ = NE_check(r_f_t_p_1,r_w_t_p_1, A, delta, n_rounds, terminals_firm, terminals_work)
                                        if is_NE:
                                            break
                                
                                if verbose:
                                    print("----initial strategies----")
                                    print(f"beta_f: {beta_f}")
                                    print(f"beta_w: {beta_w}")

                                if verbose:
                                    print("----final convergence----")
                                    print("--firm--")
                                    print(r_f_t_p_1)
                                    print("--worker--")
                                    print(r_w_t_p_1)
                                
                                # print(get_highest_prob_strat_at_level(1, terminals_firm, r_f_t_p_1))
                                r_w_rel = produce_relative_probs(r_w_t_p_1, terminals_work, False, n_rounds)
                                r_f_rel = produce_relative_probs(r_f_t_p_1, terminals_firm, True, n_rounds)
                                incredible_threats = incredible_threat_check(n_rounds, terminals_firm, r_w_rel, r_f_rel)
                                credible_threats = credible_threat_check(delta, r_f_rel ,r_w_rel, A, terminals_dict_w)
                                is_NE, strategies, relative_strategies, optimal_strategy_utils, expected_strategy_utils = NE_check(r_f_t_p_1,r_w_t_p_1, A, delta, n_rounds, terminals_firm, terminals_work)
                                file_path = f'D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}_f{tuple(n_f)}_w{tuple(n_w)}'
                                postfix=f'mixed_f({dist_f})_w({dist_w})'
                                draw_tree(n_rounds, file_path, A, r_f_rel,r_w_rel, incredible_threats, credible_threats, postfix=postfix)#, show_all_probs=False)
                                data_dict = {
                                    'initial_conditions': (prev_f[0], prev_w[0]),
                                    'distribution_params': (dist_f, dist_w),
                                    'mixed_initial_strat_params': (n_f, n_w),
                                    'final_strats': strategies, 
                                    'relative_final_strats': relative_strategies, 
                                    'optimal_strategy_utilities': optimal_strategy_utils,
                                    'expected_strategy_utilities_non_negligible': expected_strategy_utils,
                                    'incredible threats': (incredible_threats, None),
                                    'credible threats': (credible_threats, None),
                                    'NE': is_NE,
                                    'Timesteps':timesteps
                                }

                                df = pd.DataFrame(data_dict)

                                filepath = f'{n_rounds}_round_ne_convergence_files_mixed/D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_{postfix}.xlsx'
                                # fig.show()

                                try:
                                    with pd.ExcelWriter(filepath, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
                                        df.to_excel(writer, sheet_name=f'f {tuple(n_f)}, w {tuple(n_w)}', index=False)
                                except FileNotFoundError:
                                    with pd.ExcelWriter(filepath, mode='w', engine='openpyxl') as writer:
                                        df.to_excel(writer, sheet_name=f'f {tuple(n_f)}, w {tuple(n_w)}', index=False)

                                if verbose:
                                    print(f"Coverged to NE: {is_NE}")
                                    print("--------------------------")
                                    print()
    generate_imshow_from_sheet(A, pure=False, n_rounds=n_rounds, filepath=filepath, normal=False)

def run_3_round_grid_search(D, T, M, A, delta, eps, verbose=False):
    n_rounds = 3
    terminals_firm, terminals_work, _, _ = generate_terminal_seqs(A, n_rounds)
    if n_rounds % 2 == 1:
        terminal_seqs = terminals_firm
    else:
        terminal_seqs = terminals_work

    u_coeffs_worker = generate_u_coefficients(len(terminals_work), terminal_seqs, delta, worker=True)
    u_coeffs_firm = generate_u_coefficients(len(terminals_firm), terminal_seqs, delta, worker=False)

    for n_f1 in range(0,D+1):
        for n_f2 in range(0,D+1):
            for n_f3 in range(0,D+1):
                # payoffs = []
                for n_w1 in range(0,D+1):
                    for n_w2 in range(0,D+1):
                        for n_w3 in range(0,D+1):
                            print(f"Firm strategy indices {(n_f1, n_f2, n_f3)}, Worker strategy indices {(n_w1, n_w2, n_w3)}")
                            # generate random pure initial strategies
                            beta_f, beta_w, n_f, n_w = generate_init_strategies(A, n_rounds, True, [n_f1, n_f2, n_f3], [n_w1, n_w2, n_w3])
                            file_path = f'D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}_f{tuple(n_f)}_w{tuple(n_w)}'
                            
                            prev_f = np.array([beta_f])
                            prev_w = np.array([beta_w])

                            timesteps=0
                            for t in tqdm.tqdm(range(T)):
                                timesteps+=1
                                r_f_t_p_1 = agent_update_memoized(n_rounds, prev_w, A, len(terminals_firm), M, u_coeffs_firm, worker=False)
                                r_w_t_p_1 = agent_update_memoized(n_rounds, prev_f, A, len(terminals_work), M, u_coeffs_worker, worker=True)

                                prev_f = np.vstack([prev_f, r_f_t_p_1])
                                prev_w = np.vstack([prev_w, r_w_t_p_1])

                                if check_convergence(prev_f, prev_w, t, convergence_threshold=1e-7):
                                    print("Converged after", t, "steps")
                                    is_NE, _, _, _, _ = NE_check(r_f_t_p_1,r_w_t_p_1, A, delta, n_rounds, terminals_firm, terminals_work)
                                    if is_NE:
                                        break
                            
                            if verbose:
                                print("----initial strategies----")
                                print(f"beta_f: {beta_f}")
                                print(f"beta_w: {beta_w}")

                            if verbose:
                                print("----final convergence----")
                                print("--firm--")
                                print(r_f_t_p_1)
                                print("--worker--")
                                print(r_w_t_p_1)
                            
                            # print(get_highest_prob_strat_at_level(1, terminals_firm, r_f_t_p_1))
                            r_w_rel = produce_relative_probs(r_w_t_p_1, terminals_work, False, n_rounds)
                            r_f_rel = produce_relative_probs(r_f_t_p_1, terminals_firm, True, n_rounds)
                            is_NE, strategies, relative_strategies, optimal_strategy_utils, expected_strategy_utils = NE_check(r_f_t_p_1,r_w_t_p_1, A, delta, n_rounds, terminals_firm, terminals_work)
                            file_path = f'D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}_f{tuple(n_f)}_w{tuple(n_w)}'
                            postfix='grid'
                            draw_tree(n_rounds, file_path, A, r_f_rel,r_w_rel, [], [], postfix=postfix)#, show_all_probs=False)
                            data_dict = {
                                'initial_conditions': (prev_f[0], prev_w[0]),
                                'initial_strats': (n_f, n_w),
                                'final_strats': strategies, 
                                'relative_final_strats': relative_strategies, 
                                'optimal_strategy_utilities': optimal_strategy_utils,
                                'expected_strategy_utilities_non_negligible': expected_strategy_utils,
                                'NE': is_NE,
                                'Timesteps':timesteps
                            }

                            df = pd.DataFrame(data_dict)

                            filepath = f'{n_rounds}_round_ne_convergence_files_grid/D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}.xlsx'
                            # fig.show()

                            try:
                                with pd.ExcelWriter(filepath, mode='a', engine='openpyxl', if_sheet_exists='new') as writer:
                                    df.to_excel(writer, sheet_name=f'f {tuple(n_f)}, w {tuple(n_w)}', index=False)
                            except FileNotFoundError:
                                with pd.ExcelWriter(filepath, mode='w', engine='openpyxl') as writer:
                                    df.to_excel(writer, sheet_name=f'f {tuple(n_f)}, w {tuple(n_w)}', index=False)

                            if verbose:
                                print(f"Coverged to NE: {is_NE}")
                                print("--------------------------")
                                print()


def gather_3_round_data_from_sheet(filepath):
    # D=4 (['0.75000', '0.50000', '0.67500'], ['0.25000', '0.50000', '0.22500'], [((1, 2, 4), (2, 3, 0)), ((3, 4, 0), (1, 2, 2)), ((0, 0, 1), (2, 1, 2))])
    # D=5 (['0.80000', '0.60000', '0.72000'], ['0.20000', '0.40000', '0.18000'], [((4, 2, 0), (1, 2, 0)), ((2, 4, 5), (0, 2, 2)), ((1, 4, 3), (2, 0, 0))])
    excel_file = pd.ExcelFile(filepath)
    fpath, fname = filepath.split('/')

    pattern = r"D_(\d+)_M_([0-9.]+)_T_([0-9]+)_del_([0-9.]+)_eps_([0-9eE.-]+)"
    match = re.match(pattern, fname)

    D, M, T, delta, eps = match.groups()

    D = int(D)
    M = float(M)
    T = int(T)
    delta = float(delta)
    eps = float(eps)
    sheet_names = excel_file.sheet_names
    utils_f = []
    utils_w = []
    params = []
    for sheet_name in sheet_names:
        pattern = r"f\s\((.*?)\),\s*w\s\((.*?)\)"
        match = re.match(pattern, sheet_name)

        n_f_str, n_w_str = match.groups()
        n_f = tuple(map(int, n_f_str.split(',')))
        n_w = tuple(map(int, n_w_str.split(',')))

        df = excel_file.parse(sheet_name)

        f_util, w_util = df['expected_strategy_utilities_non_negligible'] 
        f_util = f'{f_util:.5f}'
        w_util = f'{w_util:.5f}'
        if w_util not in utils_w:
            utils_w.append(w_util)
            utils_f.append(f_util)
            params.append((n_f,n_w))
    return utils_f, utils_w, params


def generate_imshow_from_sheet(A, pure, n_rounds, filepath, normal=False, ne_threshold=None):
    
    excel_file = pd.ExcelFile(filepath)
    fname = filepath

    pattern = r"D_(\d+)_M_([0-9.]+)_T_([0-9]+)_del_([0-9.]+)_eps_([0-9eE.-]+)"
    match = re.match(pattern, fname)

    D, M, T, delta, eps = match.groups()
    
    D = int(D)
    M = float(M)
    T = int(T)
    delta = float(delta)
    if ne_threshold is None:
        eps = float(eps)
    else:
        eps = ne_threshold
    # print(D, M, T, delta, eps, n_rounds)

    sheet_names = excel_file.sheet_names
    # print(sheet_names)
    if n_rounds == 2:
        util_matrix = np.zeros(((D+1)**2, (D+1)**2))
    else:
        util_matrix = np.zeros(((D+1), (D+1)))
    dist_f_str = ''
    dist_w_str = ''
    if not pure:
        pattern = r"mixed_f\(\[(.*?)\]\)_w\((.*?)\)"
        # postfix=f'mixed_f({dist_f})_w({dist_w})'
        match = re.search(pattern, fname)

        dist_f_str, dist_w_str = match.groups()
        print("dist_f:", dist_f_str)
        print("dist_w:", dist_w_str)

    # Loop through each sheet and read the data
    for sheet_name in sheet_names:
        if pure:
            # Pattern to match tuples for f and w
            pattern = r"f\s\((.*?)\),\s*w\s\((.*?)\)"
            
            match = re.match(pattern, sheet_name)

            n_f_str, n_w_str = match.groups()
            if n_rounds == 1:
                n_f = int(n_f_str[:-1])
                n_w = int(n_w_str[:-1])
            else:
                n_f = tuple(map(int, n_f_str.split(',')))
                n_w = tuple(map(int, n_w_str.split(',')))

            # print(n_f, n_w)

            df = excel_file.parse(sheet_name)
        else:
            df = excel_file.parse(sheet_name)
            n_f_str, n_w_str = df['mixed_initial_strat_params'] 
            n_f_str = n_f_str.replace('[', '').replace(']','')
            n_w_str = n_w_str.replace('[', '').replace(']','')
            n_f = tuple(map(int, n_f_str.split(',')))
            n_w = tuple(map(int, n_w_str.split(',')))
            # print(n_f, n_w)
            # thresh = n_w[2]
        if n_rounds == 2:
            y_index = n_f[0] * (len(A)) + n_f[1]
            x_index = n_w[0] * (len(A)) + n_w[1]
        else:
            y_index = n_f
            x_index = n_w
        # print(y_index, x_index)
        # print(df.columns)
        actual_util = df['expected_strategy_utilities_non_negligible'] 
        is_NE = df['NE'][0]
        # print(actual_util[0])
        if is_NE or ne_threshold is not None:
            if normal:
                util_matrix[y_index, x_index] = 1-round(actual_util[0], 4)
            else:
                util_matrix[y_index, x_index] = round(actual_util[1], 4)
        else:
            util_matrix[y_index, x_index] = 1

    fig, ax = plt.subplots(figsize=(12, 10))
    if n_rounds == 1:
        pattern = r"ref\((.*?)\)"
        # postfix=f'mixed_f({dist_f})_w({dist_w})'
        match = re.search(pattern, fname)
        alpha = None
        if match:
            alpha_str = match.groups()[0]
            # print(alpha_str)
            alpha = tuple(map(int, alpha_str.split(',')))
        title = f"Worker Util Value at NE \n(η={M:.2f}, D={D}, T={T}, ɛ={eps:.0e}, reference={alpha}"
    else:
        title = f"Worker Util Value at NE \n(η={M:.2f}, D={D}, T={T}, δ={delta:.2f}, ɛ={eps:.0e}"
    if not pure:
        title+=f', f_mix={dist_f_str}, w_mix={dist_f_str})'
    else:
        title+=')'
    fig.suptitle(title)

    colors = [
        (0.0, '#9cd3f2'), # light blue
        (0.5, '#3ea1b1'), # teal
        (1.0, '#1f4e79')  # deep blue
    ]
    # colors = [
    #     (0.0, '#3D9970'), 
    #     (0.5, '#1f4e79'),
    #     (1.0, '#FF851B'), 
    # ]

    # Other color schemes: 
    # Greens: #2E8B57, #9ACD32, #98FB98
    # Sunset: #B22222, #FF7F50, #FFDAB9
    # Autumn: #8B0000, #DAA520, #CD853F
    # Retro: #3D9970, #FF851B, #FFDC00
    # Winter: #BBDEFB, #B0BEC5, #B2DFDB

    # Create the custom colormap
    # custom_cmap = LinearSegmentedColormap.from_list('custom_cmap', colors)

    im = ax.imshow(util_matrix, cmap='viridis', vmin=0, vmax=1, origin='lower')
    # im = ax.imshow(util_matrix, cmap=custom_cmap, origin='lower')

    # Add colorbar
    plt.colorbar(im, ax=ax, label='Expected Utility for Worker')

    # Customize x-axis labels
    if n_rounds == 2:
        x_labels = [f"({A[n_w1]:.2f},{A[n_w2]:.2f})" for n_w1 in range(D+1) for n_w2 in range(D+1)]
        y_labels = [f"({A[n_f1]:.2f},{A[n_f2]:.2f})" for n_f1 in range(D+1) for n_f2 in range(D+1)]
    else:
        x_labels = [f"{A[n_w1]:.2f}" for n_w1 in range(D+1)]
        y_labels = [f"{A[n_f1]:.2f}" for n_f1 in range(D+1)]
    ax.set_xticks(range(len(x_labels)))
    ax.set_xticklabels(x_labels, rotation=90, ha='center')
    ax.set_yticks(range(len(y_labels)))
    ax.set_yticklabels(y_labels)
    
    if n_rounds == 1:
        ax.set_xlabel('Worker Initial Strategy (Acceptance Threshold)')
        ax.set_ylabel('Firm Initial Strategy (Offer)')
    elif pure:
        ax.set_xlabel("Worker Initial Strategy (Acceptance Threshold, Counter Offer)")
        ax.set_ylabel("Firm Initial Strategy (Offer, Acceptance Threshold)")
    else:
        ax.set_xlabel("Worker Strategy (Low Offer, High Offer)")
        ax.set_ylabel("Firm Initial Strategy (Offer A, Offer B)")

    plt.setp(ax.get_xticklabels(), rotation=90, ha="right", rotation_mode="anchor")
    plt.tight_layout()
    # plt.show()
    plt.savefig(f'{n_rounds}_round_imshow_{fname[0:-5]}.png') 


# permutation of first two offers to mix over (firm)
# permutation of low and high offers to mix over (work)
# get from 'mixed_initial_strat_params': (n_f, n_w) 
   # but only need first two elements of tuple from n_w
# add to title: dists for both firm and worker
# ax.set_xlabel("Worker Strategy (Low Offer, High Offer)")
# ax.set_ylabel("Firm Initial Strategy (Offer A, Offer B)")

def generate_tree_from_sheet(A, n_rounds, filepath, postfix='grid'):
    excel_file = pd.ExcelFile(filepath)
    fpath, fname = filepath.split('/')

    pattern = r"D_(\d+)_M_([0-9.]+)_T_([0-9]+)_del_([0-9.]+)_eps_([0-9eE.-]+)"
    match = re.match(pattern, fname)

    D, M, T, delta, eps = match.groups()
    
    D = int(D)
    M = float(M)
    T = int(T)
    delta = float(delta)
    eps = float(eps)
    # print(D, M, T, delta, eps, n_rounds)
    

    sheet_names = excel_file.sheet_names
    # print(sheet_names)

    # Loop through each sheet and read the data
    for sheet_name in sheet_names:
        # Pattern to match tuples for f and w
        pattern = r"f\s\((.*?)\),\s*w\s\((.*?)\)"
        
        match = re.match(pattern, sheet_name)

        n_f_str, n_w_str = match.groups()
        if n_rounds == 1:
            n_f = int(n_f_str[:-1])
            n_w = int(n_w_str[:-1])
        else:
            n_f = tuple(map(int, n_f_str.split(',')))
            n_w = tuple(map(int, n_w_str.split(',')))
            # print(n_f)
            # if n_f[0] <4:
            #     continue

        if n_rounds == 2:
                y_index = n_f[0] * (len(A)) + n_f[1]
                x_index = n_w[0] * (len(A)) + n_w[1]
        else:
            y_index = n_f
            x_index = n_w
        # print(y_index, x_index)
        # print(df.columns)
        df = excel_file.parse(sheet_name, dtype=object)
        index = f'D_{D}_M_{M}_T_{T}_del_{delta}_eps_{eps:.0e}_rounds_{n_rounds}'
        file_path = index+'/'+sheet_name
        
        # print(df.columns)
        credible_threats, _ = df['credible threats']
        # print(credible_threats)
        incredible_threats, _ = df['incredible threats']
        # print(incredible_threats)
        r_f_rel, r_w_rel = df['relative_final_strats']
        credible_threats = ast.literal_eval(credible_threats)
        incredible_threats = ast.literal_eval(incredible_threats)
        r_f_rel = list(map(float, r_f_rel[1:-1].strip().split()))
        r_w_rel = list(map(float, r_w_rel[1:-1].replace('np.float64(','').replace(')','').strip().split(',')))
        # print(credible_threats)
        # print(incredible_threats)
        # print(r_f_rel)
        # print(r_w_rel)
        
        draw_tree(n_rounds, file_path, A, r_f_rel,r_w_rel, incredible_threats, credible_threats, postfix=postfix)#, show_all_probs=False)
        
def compute_meta_game_equilibrium(filepath, normal=False):
    excel_file = pd.ExcelFile(filepath)
    fpath, fname = filepath.split('/')

    pattern = r"D_(\d+)_M_([0-9.]+)_T_([0-9]+)_del_([0-9.]+)_eps_([0-9eE.-]+)_rounds_([0-9]+)"
    match = re.match(pattern, fname)
    
    D, M, T, delta, eps, n_rounds = match.groups()
    # print(n_rounds)
    D = int(D)
    M = float(M)
    T = int(T)
    n_rounds = int(n_rounds)
    A = [i / (D) for i in range(D + 1)]
    delta = float(delta)
    eps = float(eps)

    sheet_names = excel_file.sheet_names

    if n_rounds == 2:
        util_matrix = np.zeros(((D+1)**2, (D+1)**2))
    else:
        util_matrix = np.zeros(((D+1), (D+1)))

    # Loop through each sheet and read the data
    for sheet_name in sheet_names:
        # Pattern to match tuples for f and w
        pattern = r"f\s\((.*?)\),\s*w\s\((.*?)\)"
        
        match = re.match(pattern, sheet_name)

        n_f_str, n_w_str = match.groups()
        if n_rounds == 1:
            n_f = int(n_f_str[:-1])
            n_w = int(n_w_str[:-1])
            y_index = n_f
            x_index = n_w
        else:
            n_f = tuple(map(int, n_f_str.split(',')))
            n_w = tuple(map(int, n_w_str.split(',')))
            y_index = n_f[0] * (len(A)) + n_f[1]
            x_index = n_w[0] * (len(A)) + n_w[1]

        df = excel_file.parse(sheet_name)

        actual_util = df['expected_strategy_utilities_non_negligible'] 
        is_NE = df['NE'][0]

        if is_NE:
            if normal:
                util_matrix[y_index, x_index] = 1-round(actual_util[0], 4)
            else:
                util_matrix[y_index, x_index] = round(actual_util[1], 4)
        else:
            util_matrix[y_index, x_index] = 1

    if n_rounds == 1:
        A = np.array([[(1-w)-0.5 for c,w in enumerate(row)]for r,row in enumerate(util_matrix)]) # input row player values

        meta_game = nash.Game(A)

        # minimax solution
        f, w = meta_game.linear_program()# returns row strategy, col. strategy

        print("Minimax solution")
        print(f"firm equilibrium strategy: {f}")
        print(f"worker equilibrium strategy: {w}")
        print(f"Value of game: {np.transpose(f)@A@w}")

    if n_rounds == 2:
        firm_payoffs = np.array([[(1-w) for c,w in enumerate(row)]for r,row in enumerate(util_matrix)])

        meta_game = nash.Game(firm_payoffs, util_matrix)

        # support enumeration solution - finds some equilibria of degenerate games
        print("vertex enumeration")
        equilibria = meta_game.vertex_enumeration()
        for eq in equilibria:
            f,w = eq
            print(f"firm equilibrium strategy: {f}")
            print(f"worker equilibrium strategy: {w}")
            # print(f"Value of game: {np.transpose(f)@A@w}")

def get_updated_ne_threshold(filepath):
    excel_file = pd.ExcelFile(filepath)
    sheet_names = excel_file.sheet_names
    cur_threshold = 1e-7
    for sheet_name in sheet_names:
        df = excel_file.parse(sheet_name)
        is_NE = df['NE'][0]

        if not is_NE:
            actual_util = df['expected_strategy_utilities_non_negligible'] 
            optimal_util = df['optimal_strategy_utilities']
            result_tuple = tuple(abs(x - y) for x, y in zip((actual_util[0], actual_util[1]), (optimal_util[0], optimal_util[1])))
            while (result_tuple > (cur_threshold,cur_threshold)):
                cur_threshold*=10
    generate_imshow_from_sheet(A, pure=True, n_rounds=n_rounds, filepath=filepath, normal=False, ne_threshold=cur_threshold)
    return cur_threshold

if __name__ == "__main__":
    # D = 3
    D = 5 # 2 rounds
    T = 15000 # 2 rounds
    M = [0.25,0.8]
    delta = 0.1
    eps = 1e-6
    n_rounds = 2
    n_trials = 100
    A = [round(i/D,4) for i in range(0,D+1)] # action list for offers {1/D, ... , 1}
    for m in M:
        run_2_round_grid_search(D, T, m, A, delta, eps)
    # filepath = "2_round_ne_convergence_files_grid/D_5_M_0.5_T_15000_del_0.5_eps_1e-07_rounds_2.xlsx"
    # filepath= "D_30_M_0.5_T_5000_del_0.5_eps_1e-07_rounds_1.xlsx"
    # generate_imshow_from_sheet(A, True, n_rounds, filepath,normal=True)
    # generate_tree_from_sheet(A, n_rounds, filepath)
    # run_2_round_grid_search(D, T, M, A, delta, eps)
    # run_n_rounds_extended_form(D, T, M, A, eps, n_rounds, initial=[(3,4),(2,1)])
    # run_1_round_normal_bar(A, T, False, M)
    # run_n_rounds_extended_form(D, T, M, A, eps, n_rounds, initial=([1,2,3],[1,2,3]), verbose=False)
    # make_bar_graph(filepath, uniform_firm=False)
    # reference_values = [(5,15), (20,16)]
    # for ref_value in reference_values:
    #     for m in M:
    #         run_1_round_normal_grid(A, T, m, reference=ref_value, eps=1e-7)
    exit()
